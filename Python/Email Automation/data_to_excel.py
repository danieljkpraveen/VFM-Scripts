import re
import sys
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl
import datetime


def read_stdin():
    # Read all input from stdin (no prompt)
    return sys.stdin.read()


def is_html(text):
    return bool(BeautifulSoup(text, "html.parser").find())


def extract_rules(text):
    """Extract rules and their values from the input."""
    pattern = r'(Rules where .*?=)\[(.*?)\]'
    matches = re.findall(pattern, text, re.DOTALL)
    extracted = []
    for rule_desc, values in matches:
        task = rule_desc.strip().rstrip('=').strip()
        result = ', '.join([v.strip().strip('"').strip("'")
                           for v in values.split(',')])
        extracted.append([task, result])
    return extracted


def html_to_text(html):
    soup = BeautifulSoup(html, "html.parser")
    return soup.get_text(separator="\n")


def extract_interface_table_from_html(html):
    """Extracts interface table from Cisco CLI HTML output."""
    soup = BeautifulSoup(html, "html.parser")
    cli_div = soup.find("div", class_="cli-output")
    if not cli_div:
        return [], []

    lines = cli_div.get_text(separator="\n").splitlines()
    table_start = None
    for idx, line in enumerate(lines):
        if line.strip().startswith("Interface"):
            table_start = idx
            break
    if table_start is None:
        return [], []

    headers = ["Interface", "IP-Address", "Method", "Status", "Protocol"]
    data = []
    for line in lines[table_start+1:]:
        if not line.strip():
            continue
        parts = [p for p in line.replace('\xa0', ' ').split(' ') if p]
        if len(parts) < 6:
            continue
        interface = parts[0]
        ip_addr = parts[1]
        method = parts[3]
        status = parts[4]
        protocol = parts[5]
        data.append([interface, ip_addr, method, status, protocol])
    return headers, data


def text_data_to_excel(data, file_path):
    """Save plain text extracted data to Excel."""
    df = pd.DataFrame(data, columns=["Tasks", "Results"])
    df.to_excel(file_path, index=False)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for column_cells in ws.columns:
        length = max(len(str(cell.value))
                     if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    wb.save(file_path)
    print(f"Data saved to {file_path}")


def html_data_to_excel(data, file_path, headers):
    """Save HTML extracted interface table to Excel with formatting."""
    df = pd.DataFrame(data, columns=headers)
    df.to_excel(file_path, index=False)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Bold headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value))
                     if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Highlight rows based on Status
    if "Status" in headers:
        status_col = headers.index("Status") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status = row[status_col-1].value
            if status and str(status).lower() == "up":
                for cell in row:
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status and str(status).lower() == "down":
                for cell in row:
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb.save(file_path)
    print(f"Data saved to {file_path}")


def main():
    # Only read from stdin if data is piped in
    if sys.stdin.isatty():
        print(
            "Error: No input detected. Please pipe data into this script.", file=sys.stderr)
        sys.exit(1)
    input_text = sys.stdin.read()
    if not input_text.strip():
        print("Error: No input data received.", file=sys.stderr)
        sys.exit(1)
    if is_html(input_text):
        headers, table_data = extract_interface_table_from_html(input_text)
        if table_data:
            now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = f"interfaces_{now}.xlsx"
            html_data_to_excel(table_data, excel_path, headers)
            print(
                f"Extracted {len(table_data)} interfaces. Data saved to {excel_path}")
            return
        else:
            print("No interface table found in the HTML input.")
    else:
        text = input_text
        rules = extract_rules(text)
        if rules:
            now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = f"output_{now}.xlsx"
            text_data_to_excel(rules, excel_path)
            print(f"Extracted {len(rules)} rules. Data saved to {excel_path}")
        else:
            print("No rules found in the input.")


if __name__ == "__main__":
    main()
